import streamlit as st
import pandas as pd
import io
from datetime import datetime, date
from formatters import (
    apply_currency_formatting, apply_currency_or_empty_formatting,
    create_fy_highlighting_function, format_dataframe_index_as_string,
    format_currency, clean_numeric_for_display, apply_conditional_formatting
)
from calculations import (
    calculate_income_realization_metrics, calculate_monthly_realization_breakdown,
    calculate_study_pipeline_breakdown, calculate_site_realization_breakdown
)

def create_enhanced_excel_export(calendar_df, patients_df, visits_df, site_column_mapping, unique_sites, include_financial=True):
    """
    Create Excel export with enhanced explanatory headers and documentation
    
    Args:
        calendar_df: Main calendar DataFrame
        patients_df: Patients DataFrame
        visits_df: Visits DataFrame
        site_column_mapping: Site column mapping dictionary
        unique_sites: List of unique site names
        include_financial: If True, include financial columns (Income, Totals). If False, exclude them.
    
    Returns:
        BytesIO: Excel file buffer, or None if error
    """
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        st.error("openpyxl library required for enhanced Excel formatting")
        return None
    
    # Clean the dataframe for Excel export - handle pandas NA values
    def clean_for_excel(df):
        """Clean dataframe for Excel export by handling NA values and Period objects"""
        if df.empty:
            return df
            
        cleaned_df = df.copy()
        
        # Step 1: Handle Period objects FIRST (most important)
        for col in cleaned_df.columns:
            if hasattr(cleaned_df[col].dtype, 'name'):
                dtype_str = str(cleaned_df[col].dtype).lower()
                if 'period' in dtype_str:
                    # Convert Period to string representation
                    cleaned_df[col] = cleaned_df[col].astype(str)
                    continue
            
            # Check for Period objects in object dtype columns
            if cleaned_df[col].dtype == 'object' and not cleaned_df[col].empty:
                first_valid = cleaned_df[col].dropna().head(1)
                if len(first_valid) > 0:
                    if 'Period' in str(type(first_valid.iloc[0])):
                        cleaned_df[col] = cleaned_df[col].astype(str)
        
        # Step 2: Handle pandas NA values
        for col in cleaned_df.columns:
            if cleaned_df[col].dtype == 'object':
                cleaned_df[col] = cleaned_df[col].where(pd.notna(cleaned_df[col]), None)
            elif str(cleaned_df[col].dtype).startswith('Int') or str(cleaned_df[col].dtype).startswith('Float'):
                # Convert nullable integer/float to regular float
                cleaned_df[col] = cleaned_df[col].astype('float64', errors='ignore')
                cleaned_df[col] = cleaned_df[col].where(pd.notna(cleaned_df[col]), 0)
        
        # Step 3: Handle datetime columns
        for col in cleaned_df.columns:
            if pd.api.types.is_datetime64_any_dtype(cleaned_df[col]):
                # Don't convert to string here - let Excel handle it with formatting
                pass
        
        # Step 4: Final NA cleanup - replace any remaining NaN with empty string
        cleaned_df = cleaned_df.fillna('')
        
        return cleaned_df
    
    # Clean all input dataframes
    try:
        enhanced_df = clean_for_excel(calendar_df)
        clean_patients_df = clean_for_excel(patients_df) if not patients_df.empty else patients_df
        clean_visits_df = clean_for_excel(visits_df) if not visits_df.empty else visits_df
    except Exception as clean_error:
        st.error(f"Error cleaning data for Excel: {clean_error}")
        return None

    # Filter out financial columns if requested
    if not include_financial:
        # Identify financial columns to exclude
        financial_keywords = ['Income', 'Total', 'Payment', 'Revenue', 'Cost']
        columns_to_keep = []
        
        for col in enhanced_df.columns:
            # Keep non-financial columns
            is_financial = any(keyword in col for keyword in financial_keywords)
            if not is_financial:
                columns_to_keep.append(col)
            else:
                # Log which columns are being excluded
                from helpers import log_activity
                log_activity(f"Excluding financial column from Basic Excel: {col}", level='info')
        
        # Filter the dataframe
        enhanced_df = enhanced_df[columns_to_keep]
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # === Main Calendar Sheet ===
    ws_calendar = wb.active
    ws_calendar.title = "Clinical Trial Calendar"
    
    # Add title and metadata
    ws_calendar['A1'] = "Clinical Trial Calendar - Patient Visit Schedule"
    ws_calendar['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws_calendar.merge_cells('A1:H1')
    
    ws_calendar['A2'] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_calendar['A2'].font = Font(size=10, italic=True)
    
    total_patients = len(clean_patients_df) if not clean_patients_df.empty else 0
    total_sites = len(unique_sites) if unique_sites else 0
    ws_calendar['A3'] = f"Total Patients: {total_patients} | Total Sites: {total_sites}"
    ws_calendar['A3'].font = Font(size=10, italic=True)
    
    # Column explanations
    explanations = {
        'Date': 'Calendar Date (DD/MM/YYYY)',
        'Day': 'Day of Week',
        'Daily Total': 'Total Daily Revenue (£)',
        'Monthly Total': 'Cumulative Monthly Revenue (£)', 
        'FY Total': 'Cumulative Financial Year Revenue (£)'
    }
    
    # Add site grouping headers row
    site_headers_row = [''] * len(enhanced_df.columns)
    column_explanations_row = [''] * len(enhanced_df.columns)
    
    for col_idx, col_name in enumerate(enhanced_df.columns):
        # Add explanations for standard columns
        if col_name in explanations:
            column_explanations_row[col_idx] = explanations[col_name]
        
        # Add site headers for patient columns
        if col_name not in ['Date', 'Day'] and not any(x in col_name for x in ['Income', 'Total']):
            # Find which site this patient belongs to
            for site in unique_sites:
                site_data = site_column_mapping.get(site, {})
                site_columns = site_data.get('columns', [])
                if col_name in site_columns:
                    site_headers_row[col_idx] = f"Site: {site}"
                    # Enhanced patient column explanation
                    if '_' in col_name:
                        try:
                            study, patient_id = col_name.split('_', 1)
                            column_explanations_row[col_idx] = f"Study: {study} | Patient: {patient_id}"
                        except ValueError:
                            column_explanations_row[col_idx] = f"Patient: {col_name}"
                    else:
                        column_explanations_row[col_idx] = f"Patient: {col_name}"
                    break
    
    # Write to Excel starting from row 5
    start_row = 5
    
    # Site headers row
    for col_idx, header in enumerate(site_headers_row, 1):
        if header:
            cell = ws_calendar.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    # Column explanations row
    for col_idx, explanation in enumerate(column_explanations_row, 1):
        if explanation:
            cell = ws_calendar.cell(row=start_row + 1, column=col_idx, value=explanation)
            cell.font = Font(size=9, italic=True, color="595959")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Column names row
    for col_idx, col_name in enumerate(enhanced_df.columns, 1):
        cell = ws_calendar.cell(row=start_row + 2, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows - handle values carefully with UK accounting format + date-based styling
    for row_idx, (_, row) in enumerate(enhanced_df.iterrows()):
        # Get the date for this row (for styling)
        row_date = None
        if 'Date' in enhanced_df.columns:
            try:
                date_val = enhanced_df.iloc[row_idx]['Date']
                if pd.notna(date_val):
                    if isinstance(date_val, str):
                        row_date = pd.to_datetime(date_val, format='%d/%m/%Y', errors='coerce')
                    else:
                        row_date = pd.to_datetime(date_val)
            except:
                row_date = None
        
        # Determine row styling based on date
        # NOTE: We do NOT highlight today's date in Excel (only on screen)
        row_fill = None
        if row_date is not None and pd.notna(row_date):
            # Check for special dates (in priority order)
            if row_date.month == 3 and row_date.day == 31:
                # Financial year end - dark blue background
                row_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            elif row_date == row_date + pd.offsets.MonthEnd(0):
                # Month end - light blue background
                row_fill = PatternFill(start_color="60A5FA", end_color="60A5FA", fill_type="solid")
            elif row_date.weekday() in (5, 6):
                # Weekend - gray background
                row_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Process each cell in the row
        for col_idx, value in enumerate(row, 1):
            # Safely convert value for Excel
            excel_value = value
            if pd.isna(value) or value is None:
                excel_value = ""
            elif hasattr(value, '__iter__') and not isinstance(value, str):
                # Handle any remaining complex objects
                excel_value = str(value)
            
            cell = ws_calendar.cell(row=start_row + 3 + row_idx, column=col_idx, value=excel_value)
            
            # Apply row-based styling (date highlighting)
            if row_fill is not None:
                cell.fill = row_fill
                # Adjust font color for dark backgrounds (only Financial Year End is dark blue)
                if row_fill.start_color == "1E40AF":
                    cell.font = Font(color="FFFFFF")
            
            # Format dates
            if col_idx == 1 and isinstance(excel_value, str) and len(str(excel_value)) > 0:  # Date column
                cell.number_format = 'DD/MM/YYYY'
            # Format currency columns with UK accounting format (only if financial columns included)
            elif include_financial and col_idx > 2 and any(x in enhanced_df.columns[col_idx-1] for x in ['Total', 'Income']):
                if isinstance(excel_value, str) and '£' in excel_value:
                    # Remove £ symbol and convert to number for proper accounting format
                    try:
                        numeric_value = float(excel_value.replace('£', '').replace(',', ''))
                        cell.value = numeric_value
                        # Simplified UK currency format
                        cell.number_format = '£#,##0.00;[Red](£#,##0.00)'
                    except (ValueError, AttributeError):
                        # Fallback to standard currency if conversion fails
                        cell.number_format = '"£"#,##0.00'
                elif isinstance(excel_value, (int, float)) and excel_value != 0:
                    # Direct numeric values
                    cell.number_format = '£#,##0.00;[Red](£#,##0.00)'
                elif excel_value == 0 or excel_value == "":
                    # Zero values - set to 0 and apply accounting format
                    cell.value = 0
                    cell.number_format = '£#,##0.00;[Red](£#,##0.00)'
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(enhanced_df.columns, 1):
        col_letter = get_column_letter(col_idx)
        try:
            max_length = max(
                len(str(site_headers_row[col_idx-1])) if col_idx <= len(site_headers_row) else 0,
                len(str(column_explanations_row[col_idx-1])) if col_idx <= len(column_explanations_row) else 0,
                len(str(col)),
                max([len(str(cell)) for cell in enhanced_df[col].astype(str)]) if not enhanced_df.empty else 0
            )
            ws_calendar.column_dimensions[col_letter].width = min(max(12, max_length + 2), 25)
        except:
            ws_calendar.column_dimensions[col_letter].width = 15  # Default width
    
    # Set row heights for header rows
    ws_calendar.row_dimensions[start_row].height = 25
    ws_calendar.row_dimensions[start_row + 1].height = 35
    ws_calendar.row_dimensions[start_row + 2].height = 20
    
    # === Data Dictionary Sheet ===
    ws_dict = wb.create_sheet("Data Dictionary")
    
    # Create dictionary data based on whether financial columns are included
    if include_financial:
        dictionary_data = [
            ["Column Name", "Description", "Data Type", "Example"],
            ["Date", "Calendar date for the visit schedule", "Date", "15/09/2025"],
            ["Day", "Day of the week", "Text", "Monday"],
            ["Daily Total", "Total revenue for all visits on this date", "Currency", "£1,250.00"],
            ["Monthly Total", "Cumulative revenue for the month up to this date", "Currency", "£15,750.00"],
            ["FY Total", "Cumulative revenue for financial year up to this date", "Currency", "£125,500.00"],
            ["Study_PatientID", "Visit information for specific patient", "Text", "V1, V2"],
            ["Site Income", "Revenue generated by visits at specific site", "Currency", "£500.00"],
            ["", "", "", ""],
            ["Special Date Highlighting", "", "", ""],
            ["Financial Year End", "Highlighted in dark blue (31 March)", "Formatting", "31/03/2025"],
            ["Month End", "Highlighted in light blue", "Formatting", "Last day of month"],
            ["Weekends", "Highlighted in gray (Saturday & Sunday)", "Formatting", "Sat/Sun"]
        ]
    else:
        dictionary_data = [
            ["Column Name", "Description", "Data Type", "Example"],
            ["Date", "Calendar date for the visit schedule", "Date", "15/09/2025"],
            ["Day", "Day of the week", "Text", "Monday"],
            ["Study_PatientID", "Visit information for specific patient", "Text", "V1, V2"],
            ["", "", "", ""],
            ["Note", "Financial columns excluded from this export", "", ""],
            ["", "", "", ""],
            ["Special Date Highlighting", "", "", ""],
            ["Financial Year End", "Highlighted in dark blue (31 March)", "Formatting", "31/03/2025"],
            ["Month End", "Highlighted in light blue", "Formatting", "Last day of month"],
            ["Weekends", "Highlighted in gray (Saturday & Sunday)", "Formatting", "Sat/Sun"]
        ]
    
    for row_idx, row_data in enumerate(dictionary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_dict.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            if row_idx == 1:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust dictionary columns
    for col_idx in range(1, 5):
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:  # Description column
            ws_dict.column_dimensions[col_letter].width = 40
        else:
            ws_dict.column_dimensions[col_letter].width = 15
    
    # === Summary Sheet ===
    ws_summary = wb.create_sheet("Summary")
    
    # Format date range safely
    date_range_text = "N/A"
    if not enhanced_df.empty and 'Date' in enhanced_df.columns:
        try:
            min_date = enhanced_df['Date'].min()
            max_date = enhanced_df['Date'].max()
            if pd.notna(min_date) and pd.notna(max_date):
                date_range_text = f"{min_date} to {max_date}"
        except:
            date_range_text = "Date range unavailable"
    
    summary_data = [
        ["Clinical Trial Summary", ""],
        ["", ""],
        ["Total Patients", total_patients],
        ["Total Sites", total_sites],
        ["Date Range", date_range_text],
        ["", ""],
        ["Sites Included:", ""],
    ]
    
    # Add site details safely
    for site in sorted(unique_sites) if unique_sites else []:
        try:
            site_data = site_column_mapping.get(site, {})
            site_columns = site_data.get('columns', [])
            site_patients = len([col for col in enhanced_df.columns if col in site_columns]) if not enhanced_df.empty else 0
            summary_data.append([f"  - {site}", f"{site_patients} patients"])
        except:
            summary_data.append([f"  - {site}", "Data unavailable"])
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=str(label) if label is not None else "").font = Font(bold=True if value == "" else False)
        ws_summary.cell(row=row_idx, column=2, value=str(value) if value is not None else "")
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    
    # === By Study Income (FY) Sheet (financial export only) ===
    try:
        if include_financial and visits_df is not None and not visits_df.empty:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            from calculations import calculate_study_realization_by_study
            
            # Defensive: ensure a 'Date' column exists (fallback from ActualDate if miswired)
            source_visits_df = visits_df
            try:
                if 'Date' not in source_visits_df.columns and 'ActualDate' in source_visits_df.columns:
                    source_visits_df = source_visits_df.copy()
                    source_visits_df['Date'] = pd.to_datetime(source_visits_df['ActualDate']).dt.normalize()
            except Exception:
                pass
            
            by_study_df = calculate_study_realization_by_study(source_visits_df, period='current_fy')
            ws_by = wb.create_sheet("By Study Income (FY)")
            
            # Title
            ws_by['A1'] = "By Study Income (Current Financial Year)"
            ws_by['A1'].font = Font(size=14, bold=True, color="1F4E79")
            
            # Headers
            headers = list(by_study_df.columns) if not by_study_df.empty else [
                'Study', 'Completed Income', 'Completed Visits', 'Scheduled Income',
                'Scheduled Visits', 'Pipeline Income', 'Remaining Visits', 'Realization Rate'
            ]
            for col_idx, col_name in enumerate(headers, 1):
                cell = ws_by.cell(row=3, column=col_idx, value=str(col_name))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            currency_cols = {'Completed Income', 'Scheduled Income', 'Pipeline Income'}
            integer_cols = {'Completed Visits', 'Scheduled Visits', 'Remaining Visits'}
            percent_cols = {'Realization Rate'}
            
            for row_idx, row in enumerate(by_study_df.itertuples(index=False), start=4):
                for col_idx, col_name in enumerate(headers, 1):
                    val = getattr(row, col_name.replace(' ', '_')) if hasattr(row, col_name.replace(' ', '_')) else getattr(row, col_name, '')
                    cell = ws_by.cell(row=row_idx, column=col_idx, value=val)
                    # Number formats
                    if col_name in currency_cols:
                        try:
                            numeric_val = float(val)
                            cell.value = numeric_val
                            cell.number_format = '£#,##0.00;[Red](£#,##0.00)'
                        except Exception:
                            pass
                    elif col_name in integer_cols:
                        try:
                            cell.value = int(val)
                            cell.number_format = '0'
                        except Exception:
                            pass
                    elif col_name in percent_cols:
                        try:
                            # Expecting percentage as 0-100
                            pct = float(val) / 100.0 if float(val) > 1 else float(val)
                            cell.value = pct
                            cell.number_format = '0.0%'
                        except Exception:
                            pass
            
            # Auto width
            for idx, col_name in enumerate(headers, 1):
                col_letter = get_column_letter(idx)
                ws_by.column_dimensions[col_letter].width = min(max(12, len(str(col_name)) + 6), 28)
    except Exception as _e:
        # Non-fatal: keep export even if this sheet fails
        pass
    try:
        wb.save(output)
        output.seek(0)
        return output
    except Exception as save_error:
        st.error(f"Error saving Excel file: {save_error}")
        return None

# Add the missing table builder functions that display_components expects
def display_income_table_pair(financial_df):
    """Display monthly income analysis tables"""
    try:
        if financial_df.empty:
            st.info("No financial data available")
            return
            
        # Fix: Move fillna operation BEFORE groupby
        financial_df['Payment'] = financial_df['Payment'].fillna(0)
        monthly_totals = financial_df.groupby('MonthYear')['Payment'].sum()
        
        # Simplify: Handle only Series case since groupby should always return a Series
        if not isinstance(monthly_totals, pd.Series):
            st.error("Unexpected data format in monthly analysis")
            return
        
        if monthly_totals.empty:
            st.info("No monthly data available")
            return
        
        # Convert to DataFrame for display
        monthly_df = monthly_totals.reset_index()
        monthly_df.columns = ['Month', 'Total Income']
        
        # Convert Period objects to strings for display
        monthly_df['Month'] = monthly_df['Month'].astype(str)
        monthly_df['Total Income'] = monthly_df['Total Income'].apply(format_currency)
        
        # Sort by month (convert to datetime for proper sorting)
        monthly_df['SortDate'] = pd.to_datetime(monthly_df['Month'])
        monthly_df = monthly_df.sort_values('SortDate')
        monthly_df = monthly_df.drop('SortDate', axis=1)
        
        st.dataframe(monthly_df, width='stretch')
    except Exception as e:
        st.error(f"Error displaying monthly income: {e}")

def display_profit_sharing_table(quarterly_ratios):
    """Display profit sharing analysis table"""
    try:
        if quarterly_ratios:
            df = pd.DataFrame(quarterly_ratios)
            # Apply highlighting for Financial Year rows
            styled_df = df.style.apply(
                lambda x: ['background-color: #e6f3ff; font-weight: bold;' if x['Type'] == 'Financial Year' else '' for _ in x], 
                axis=1
            )
            st.dataframe(styled_df, width='stretch', hide_index=True)
        else:
            st.info("No quarterly data available for profit sharing analysis")
    except Exception as e:
        st.error(f"Error displaying profit sharing table: {e}")

def display_ratio_breakdown_table(ratio_data, title):
    """Display ratio breakdown table"""
    try:
        if ratio_data:
            st.write(f"**{title}**")
            df = pd.DataFrame(ratio_data)
            st.dataframe(df, width='stretch', hide_index=True)
        else:
            st.info(f"No data available for {title}")
    except Exception as e:
        st.error(f"Error displaying {title}: {e}")

def create_summary_metrics_row(metrics_data, columns=4):
    """Create a row of metrics using Streamlit columns"""
    try:
        cols = st.columns(columns)
        for i, (label, value) in enumerate(metrics_data.items()):
            with cols[i % columns]:
                st.metric(label, value)
    except Exception as e:
        st.error(f"Error creating metrics row: {e}")

def display_breakdown_by_study(site_visits, site_patients, site_name):
    """Display study breakdown for a site"""
    try:
        study_breakdown = site_patients.groupby('Study').agg({
            'PatientID': 'count'
        }).rename(columns={'PatientID': 'Patient Count'})
        
        visit_breakdown = site_visits.groupby('Study').agg({
            'Visit': 'count',
            'Payment': 'sum'
        }).rename(columns={'Visit': 'Visit Count', 'Payment': 'Total Income'})
        
        combined_breakdown = study_breakdown.join(visit_breakdown, how='left').fillna(0)
        combined_breakdown['Total Income'] = combined_breakdown['Total Income'].apply(format_currency)
        
        st.dataframe(combined_breakdown, width='stretch')
    except Exception as e:
        st.error(f"Error displaying study breakdown: {e}")

def create_time_period_config():
    """Create time period configuration dictionary"""
    return {
        'monthly': {'column': 'MonthYear', 'name': 'Month', 'title': 'Monthly Ratio Breakdown'},
        'quarterly': {'column': 'QuarterYear', 'name': 'Quarter', 'title': 'Quarterly Ratio Breakdown'},
        'yearly': {'column': 'FinancialYear', 'name': 'Financial Year', 'title': 'Financial Year Ratio Breakdown'}
    }

def display_site_time_analysis(site_visits, site_patients, site_name, enhanced_visits_df):
    """Display time-based analysis for a site"""
    try:
        st.write("**Time-based Analysis**")
        
        # Quarterly analysis
        if 'QuarterYear' in enhanced_visits_df.columns:
            quarterly_stats = site_visits.groupby(enhanced_visits_df['QuarterYear']).agg({
                'Visit': 'count',
                'Payment': 'sum'
            }).rename(columns={'Visit': 'Visit Count', 'Payment': 'Income'})
            
            if not quarterly_stats.empty:
                quarterly_display = quarterly_stats.copy()
                quarterly_display['Income'] = quarterly_display['Income'].apply(format_currency)
                st.write("*Quarterly Summary*")
                st.dataframe(quarterly_display, width='stretch')
    except Exception as e:
        st.error(f"Error displaying time analysis: {e}")

def display_complete_realization_analysis(visits_df, trials_df, patients_df):
    """Display complete income realization analysis"""
    try:
        st.subheader("Income Realization Analysis")
        
        # Calculate metrics
        metrics = calculate_income_realization_metrics(visits_df, trials_df, patients_df)
        
        # Display summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Completed Income", format_currency(metrics['completed_income']))
        with col2:
            st.metric("Total Scheduled", format_currency(metrics['total_scheduled_income']))
        with col3:
            st.metric("Pipeline Remaining", format_currency(metrics['pipeline_income']))
        with col4:
            st.metric("Realization Rate", f"{metrics['realization_rate']:.1f}%")
        
        # Monthly breakdown
        monthly_data = calculate_monthly_realization_breakdown(visits_df, trials_df)
        if monthly_data:
            st.write("**Monthly Realization Breakdown**")
            monthly_df = pd.DataFrame(monthly_data)
            monthly_df['Completed_Income'] = monthly_df['Completed_Income'].apply(format_currency)
            monthly_df['Scheduled_Income'] = monthly_df['Scheduled_Income'].apply(format_currency)
            monthly_df['Realization_Rate'] = monthly_df['Realization_Rate'].apply(lambda x: f"{x:.1f}%")
            st.dataframe(monthly_df, width='stretch', hide_index=True)
        
        # Study pipeline breakdown
        study_pipeline = calculate_study_pipeline_breakdown(visits_df, trials_df)
        if not study_pipeline.empty:
            st.write("**Pipeline by Study**")
            study_display = study_pipeline.copy()
            study_display['Pipeline_Value'] = study_display['Pipeline_Value'].apply(format_currency)
            st.dataframe(study_display, width='stretch', hide_index=True)
        
        # Site realization breakdown
        site_data = calculate_site_realization_breakdown(visits_df, trials_df)
        if site_data:
            st.write("**Site Realization Summary**")
            site_df = pd.DataFrame(site_data)
            site_df['Completed_Income'] = site_df['Completed_Income'].apply(format_currency)
            site_df['Total_Scheduled_Income'] = site_df['Total_Scheduled_Income'].apply(format_currency)
            site_df['Pipeline_Income'] = site_df['Pipeline_Income'].apply(format_currency)
            site_df['Realization_Rate'] = site_df['Realization_Rate'].apply(lambda x: f"{x:.1f}%")
            st.dataframe(site_df, width='stretch', hide_index=True)
            
    except Exception as e:
        st.error(f"Error in realization analysis: {e}")

def display_site_screen_failures(site_patients, screen_failures):
    """Display screen failures for a site"""
    try:
        site_screen_failures = []
        for patient in site_patients.itertuples():
            patient_study_key = f"{patient.PatientID}_{patient.Study}"
            if patient_study_key in screen_failures:
                site_screen_failures.append({
                    'Patient': patient.PatientID,
                    'Study': patient.Study,
                    'Screen Fail Date': screen_failures[patient_study_key].strftime('%Y-%m-%d')
                })
        
        if site_screen_failures:
            st.write("**Screen Failures**")
            st.dataframe(pd.DataFrame(site_screen_failures), width='stretch', hide_index=True)
    except Exception as e:
        st.error(f"Error displaying screen failures: {e}")

def create_excel_export_data(calendar_df, site_column_mapping, unique_sites):
    """Create Excel export data with proper formatting"""
    try:
        excel_financial_cols = ["Daily Total", "Monthly Total", "FY Total"] + [c for c in calendar_df.columns if "Income" in c]
        excel_full_df = calendar_df.copy()
        
        # Format dates
        excel_full_df["Date"] = excel_full_df["Date"].dt.strftime("%d/%m/%Y")
        
        # Format currency columns
        for col in excel_financial_cols:
            if col in excel_full_df.columns:
                if col in ["Monthly Total", "FY Total"]:
                    excel_full_df[col] = excel_full_df[col].apply(
                        lambda v: f"£{v:,.2f}" if pd.notna(v) and v != 0 else ""
                    )
                else:
                    excel_full_df[col] = excel_full_df[col].apply(
                        lambda v: f"£{v:,.2f}" if pd.notna(v) else "£0.00"
                    )
        
        return excel_full_df
    except Exception as e:
        st.error(f"Error creating Excel export data: {e}")
        return calendar_df

def apply_excel_formatting(ws, excel_df, site_column_mapping, unique_sites):
    """Apply Excel formatting"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # Add site headers
        for col_idx, col_name in enumerate(excel_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col_name not in ["Date", "Day"] and not any(x in col_name for x in ["Income", "Total"]):
                for site in unique_sites:
                    if col_name in site_column_mapping.get(site, {}).get('columns', []):
                        ws[f"{col_letter}1"] = site
                        ws[f"{col_letter}1"].font = Font(bold=True, size=12)
                        ws[f"{col_letter}1"].fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
                        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")
                        break

        # Auto-adjust column widths
        for idx, col in enumerate(excel_df.columns, 1):
            col_letter = get_column_letter(idx)
            max_length = max([len(str(cell)) if cell is not None else 0 for cell in excel_df[col].tolist()] + [len(col)])
            ws.column_dimensions[col_letter].width = max(10, max_length + 2)
    except ImportError:
        pass  # Skip formatting if openpyxl styles not available

def add_enhanced_download_section(calendar_df, patients_df, visits_df, site_column_mapping, unique_sites):
    """Add enhanced download section with explanatory Excel export"""
    
    st.subheader("Enhanced Downloads")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Standard CSV download
        csv_data = calendar_df.to_csv(index=False)
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"clinical_calendar_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
    
    with col2:
        # Basic Excel download
        basic_excel = io.BytesIO()
        with pd.ExcelWriter(basic_excel, engine='openpyxl') as writer:
            calendar_df.to_excel(writer, sheet_name='Calendar', index=False)
        basic_excel.seek(0)
        
        st.download_button(
            label="Download Basic Excel",
            data=basic_excel.getvalue(),
            file_name=f"clinical_calendar_basic_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col3:
        # Enhanced Excel download
        enhanced_excel = create_enhanced_excel_export(
            calendar_df, patients_df, visits_df, site_column_mapping, unique_sites
        )
        
        if enhanced_excel:
            st.download_button(
                label="Download Enhanced Excel",
                data=enhanced_excel.getvalue(),
                file_name=f"clinical_calendar_enhanced_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Includes explanatory headers, data dictionary, and summary sheet"
            )
