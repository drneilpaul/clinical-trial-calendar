import streamlit as st
import pandas as pd
import io
from datetime import date
import re
import streamlit.components.v1 as components
from helpers import load_file, normalize_columns, parse_dates_column
from processing_calendar import build_calendar

# Initialize session state
if 'show_patient_form' not in st.session_state:
    st.session_state.show_patient_form = False
if 'show_visit_form' not in st.session_state:
    st.session_state.show_visit_form = False
if 'patient_added' not in st.session_state:
    st.session_state.patient_added = False
if 'visit_added' not in st.session_state:
    st.session_state.visit_added = False
if 'list_weight' not in st.session_state:
    st.session_state.list_weight = 35
if 'work_weight' not in st.session_state:
    st.session_state.work_weight = 35
if 'recruitment_weight' not in st.session_state:
    st.session_state.recruitment_weight = 30

def show_legend(actual_visits_df):
    if actual_visits_df is not None:
        st.info("""
        **Legend with Color Coding:**
        
        **Actual Visits:**
        - ✅ Visit X (Green background) = Completed Visit (within tolerance window)  
        - ⚠️ Visit X (Yellow background) = Completed Visit (outside tolerance window)
        - ❌ Screen Fail X (Red background) = Screen failure (no future visits)
        
        **Scheduled Visits:**
        - Visit X (Gray background) = Scheduled/Planned Visit
        - \\- (Light blue-gray, italic) = Before tolerance period
        - \\+ (Light blue-gray, italic) = After tolerance period
        
        **Date Formatting:**
        - Red background = Today's date
        - Light blue background = Month end
        - Dark blue background = Financial year end (31 March)
        - Gray background = Weekend
        """)
    else:
        st.info("""
        **Legend:** 
        - Visit X (Gray) = Scheduled Visit
        - - (Light blue-gray) = Before tolerance period
        - + (Light blue-gray) = After tolerance period
        """)

def display_styled_calendar(calendar_df, site_column_mapping, unique_sites):
    st.subheader("Generated Visit Calendar")
    
    final_ordered_columns = ["Date", "Day"]
    for site in unique_sites:
        site_columns = site_column_mapping.get(site, [])
        for col in site_columns:
            if col in calendar_df.columns:
                final_ordered_columns.append(col)
    
    display_df = calendar_df[final_ordered_columns].copy()
    display_df_for_view = display_df.copy()
    display_df_for_view["Date"] = display_df_for_view["Date"].dt.strftime("%Y-%m-%d")
    
    # Create site header row
    site_header_row = {}
    for col in display_df_for_view.columns:
        if col in ["Date", "Day"]:
            site_header_row[col] = ""
        else:
            site_found = ""
            for site in unique_sites:
                if col in site_column_mapping.get(site, []):
                    site_found = site
                    break
            site_header_row[col] = site_found
    
    site_header_df = pd.DataFrame([site_header_row])
    display_with_header = pd.concat([site_header_df, display_df_for_view], ignore_index=True)
    
    # Styling function
    def highlight_calendar(row):
        if row.name == 0:
            styles = []
            for col_name in row.index:
                if row[col_name] != "":
                    styles.append('background-color: #e6f3ff; font-weight: bold; text-align: center; border: 1px solid #ccc;')
                else:
                    styles.append('background-color: #f8f9fa; border: 1px solid #ccc;')
            return styles
        else:
            styles = []
            date_str = row.get("Date", "")
            date_obj = None
            try:
                if date_str:
                    date_obj = pd.to_datetime(date_str)
            except:
                pass

            today = pd.to_datetime(date.today())
            
            for col_name, cell_value in row.items():
                style = ""
                
                if date_obj is not None and not pd.isna(date_obj):
                    if date_obj.date() == today.date():
                        style = 'background-color: #dc2626; color: white; font-weight: bold;'
                    elif date_obj.month == 3 and date_obj.day == 31:
                        style = 'background-color: #1e40af; color: white; font-weight: bold;'
                    elif date_obj == date_obj + pd.offsets.MonthEnd(0):
                        style = 'background-color: #60a5fa; color: white;'
                    elif date_obj.weekday() in (5, 6):
                        style = 'background-color: #e5e7eb;'
                
                if style == "" and col_name not in ["Date", "Day"] and str(cell_value) != "":
                    cell_str = str(cell_value)
                    if "✅ Visit" in cell_str:
                        style = 'background-color: #d4edda; color: #155724; font-weight: bold;'
                    elif "⚠️ Visit" in cell_str:
                        style = 'background-color: #fff3cd; color: #856404; font-weight: bold;'
                    elif "❌ Screen Fail" in cell_str:
                        style = 'background-color: #f8d7da; color: #721c24; font-weight: bold;'
                    elif "Visit " in cell_str and not cell_str.startswith(("✅", "⚠️")):
                        style = 'background-color: #e2e3e5; color: #383d41;'
                    elif cell_str in ["+", "-"]:
                        style = 'background-color: #f1f5f9; color: #64748b; font-style: italic; font-size: 0.9em;'
                
                styles.append(style)
            
            return styles

    try:
        styled_df = display_with_header.style.apply(highlight_calendar, axis=1)
        html_table = styled_df.to_html(escape=False)
        components.html(f"""
        <div style='max-height: 700px; overflow: auto; border: 1px solid #ddd;'>
            {html_table}
        </div>
        """, height=720, scrolling=True)
    except Exception as e:
        st.write(f"Styling error: {e}")
        st.dataframe(display_with_header, use_container_width=True)

def display_financial_analysis(stats, visits_df):
    st.subheader("💰 Financial Analysis")
    
    financial_df = visits_df[
        (visits_df['Visit'].str.startswith("✅")) |
        (visits_df['Visit'].str.startswith("❌ Screen Fail")) |
        (visits_df['Visit'].str.contains('Visit', na=False) & (~visits_df.get('IsActual', False)))
    ].copy()
    
    if not financial_df.empty:
        actual_financial = financial_df[financial_df.get('IsActual', False)]
        scheduled_financial = financial_df[~financial_df.get('IsActual', True)]
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            actual_income = actual_financial['Payment'].sum() if not actual_financial.empty else 0
            st.metric("Actual Income", f"£{actual_income:,.2f}")
        with col2:
            scheduled_income = scheduled_financial['Payment'].sum() if not scheduled_financial.empty else 0
            st.metric("Scheduled Income", f"£{scheduled_income:,.2f}")
        with col3:
            total_income = actual_income + scheduled_income
            st.metric("Total Income", f"£{total_income:,.2f}")
        with col4:
            screen_fail_count = len(actual_financial[actual_financial.get('IsScreenFail', False)]) if not actual_financial.empty else 0
            st.metric("Screen Failures", screen_fail_count)

        # Monthly analysis
        financial_df['MonthYear'] = financial_df['Date'].dt.to_period('M')
        financial_df['Quarter'] = financial_df['Date'].dt.quarter
        financial_df['Year'] = financial_df['Date'].dt.year
        financial_df['QuarterYear'] = financial_df['Year'].astype(str) + '-Q' + financial_df['Quarter'].astype(str)
        financial_df['FinancialYear'] = financial_df['Date'].apply(
            lambda d: f"{d.year}-{d.year+1}" if d.month >= 4 else f"{d.year-1}-{d.year}"
        )
        
        monthly_income = financial_df.groupby(['SiteofVisit', 'MonthYear'])['Payment'].sum().reset_index()
        if not monthly_income.empty:
            monthly_pivot = monthly_income.pivot(index='MonthYear', columns='SiteofVisit', values='Payment').fillna(0)
            
            st.subheader("📊 Monthly Income Chart")
            monthly_pivot.index = monthly_pivot.index.astype(str)
            st.bar_chart(monthly_pivot)
            
            # Display financial tables
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Monthly Income by Site**")
                monthly_display = monthly_pivot.copy()
                for col in monthly_display.columns:
                    monthly_display[col] = monthly_display[col].apply(lambda x: f"£{x:,.2f}" if x != 0 else "£0.00")
                st.dataframe(monthly_display, use_container_width=True)
    
    else:
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Visits", stats.get("total_visits", 0))
        with col2:
            st.metric("Total Income", f"£{stats.get('total_income', 0):,.2f}")

def display_quarterly_profit_sharing(financial_df, patients_df):
    """Display quarterly profit sharing analysis with adjustable weights"""
    st.subheader("📊 Quarterly Profit Sharing Analysis")
    
    # Weighting adjustment button
    if st.button("⚙️ Adjust Profit Sharing Weights"):
        st.session_state.show_weights_form = True
    
    # Modal for weight adjustment
    if st.session_state.get('show_weights_form', False):
        try:
            @st.dialog("Adjust Profit Sharing Weights")
            def weights_adjustment_form():
                st.write("**Current Formula:** List Sizes + Work Done + Patient Recruitment = 100%")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write("**List Sizes Weight**")
                    list_input = st.number_input(
                        "List Size %:",
                        min_value=0,
                        max_value=100,
                        value=st.session_state.list_weight,
                        step=1
                    )
                
                with col2:
                    st.write("**Work Done Weight**")
                    work_input = st.number_input(
                        "Work Done %:",
                        min_value=0,
                        max_value=100,
                        value=st.session_state.work_weight,
                        step=1
                    )
                
                with col3:
                    st.write("**Patient Recruitment Weight**")
                    recruitment_input = st.number_input(
                        "Recruitment %:",
                        min_value=0,
                        max_value=100,
                        value=st.session_state.recruitment_weight,
                        step=1
                    )
                
                total_weight = list_input + work_input + recruitment_input
                
                if total_weight == 100:
                    st.success(f"✅ Total: {total_weight}% (Perfect!)")
                else:
                    st.error(f"❌ Total: {total_weight}% (Must equal 100%)")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("✅ Apply Changes", disabled=(total_weight != 100), use_container_width=True):
                        st.session_state.list_weight = list_input
                        st.session_state.work_weight = work_input
                        st.session_state.recruitment_weight = recruitment_input
                        st.session_state.show_weights_form = False
                        st.success("Weights updated!")
                        st.rerun()
                
                with col2:
                    if st.button("❌ Cancel", use_container_width=True):
                        st.session_state.show_weights_form = False
                        st.rerun()
            
            weights_adjustment_form()
        except AttributeError:
            st.error("Modal dialogs require Streamlit version 1.28+")
            st.session_state.show_weights_form = False
    
    # Display current weights and analysis
    st.info(f"**Current Weights:** List Sizes {st.session_state.list_weight}% • Work Done {st.session_state.work_weight}% • Patient Recruitment {st.session_state.recruitment_weight}%")
    
    # Calculate profit sharing if we have quarterly data
    if 'QuarterYear' in financial_df.columns and not financial_df.empty:
        quarters = sorted(financial_df['QuarterYear'].unique())
        
        if len(quarters) > 0:
            # Fixed list sizes
            ashfields_list_size = 28500
            kiltearn_list_size = 12500
            total_list_size = ashfields_list_size + kiltearn_list_size
            ashfields_list_ratio = ashfields_list_size / total_list_size
            kiltearn_list_ratio = kiltearn_list_size / total_list_size
            
            list_weight = st.session_state.list_weight / 100
            work_weight = st.session_state.work_weight / 100
            recruitment_weight = st.session_state.recruitment_weight / 100
            
            quarterly_data = []
            
            for quarter in quarters:
                quarter_data = financial_df[financial_df['QuarterYear'] == quarter]
                
                if len(quarter_data) == 0:
                    continue
                
                # Work ratios
                quarter_site_work = quarter_data.groupby('SiteofVisit').size()
                quarter_total_work = quarter_site_work.sum()
                
                ashfields_work_ratio = quarter_site_work.get('Ashfields', 0) / quarter_total_work if quarter_total_work > 0 else 0
                kiltearn_work_ratio = quarter_site_work.get('Kiltearn', 0) / quarter_total_work if quarter_total_work > 0 else 0
                
                # Recruitment ratios
                quarter_recruitment = quarter_data.groupby('PatientOrigin').agg({'PatientID': 'nunique'})
                quarter_total_patients = quarter_recruitment['PatientID'].sum()
                
                ashfields_recruitment_ratio = quarter_recruitment.loc['Ashfields', 'PatientID'] / quarter_total_patients if 'Ashfields' in quarter_recruitment.index and quarter_total_patients > 0 else 0
                kiltearn_recruitment_ratio = quarter_recruitment.loc['Kiltearn', 'PatientID'] / quarter_total_patients if 'Kiltearn' in quarter_recruitment.index and quarter_total_patients > 0 else 0
                
                # Calculate weighted ratios
                ashfields_final_ratio = (ashfields_list_ratio * list_weight + 
                                       ashfields_work_ratio * work_weight + 
                                       ashfields_recruitment_ratio * recruitment_weight)
                
                kiltearn_final_ratio = (kiltearn_list_ratio * list_weight + 
                                      kiltearn_work_ratio * work_weight + 
                                      kiltearn_recruitment_ratio * recruitment_weight)
                
                # Normalize
                total_ratio = ashfields_final_ratio + kiltearn_final_ratio
                if total_ratio > 0:
                    ashfields_final_ratio /= total_ratio
                    kiltearn_final_ratio /= total_ratio
                
                quarter_income = quarter_data['Payment'].sum()
                
                quarterly_data.append({
                    'Quarter': quarter,
                    'Ashfields Share': f"{ashfields_final_ratio:.1%}",
                    'Kiltearn Share': f"{kiltearn_final_ratio:.1%}",
                    'Total Income': f"£{quarter_income:,.2f}",
                    'Ashfields Income': f"£{quarter_income * ashfields_final_ratio:,.2f}",
                    'Kiltearn Income': f"£{quarter_income * kiltearn_final_ratio:,.2f}"
                })
            
            if quarterly_data:
                quarterly_df = pd.DataFrame(quarterly_data)
                st.dataframe(quarterly_df, use_container_width=True)

def patient_entry_modal():
    """Modal for adding patients"""
    @st.dialog("Add New Patient")
    def patient_form():
        patients_file = st.session_state.get('patients_file')
        trials_file = st.session_state.get('trials_file')
        
        if not patients_file or not trials_file:
            st.error("Files not available")
            return
        
        existing_patients = load_file(patients_file)
        existing_patients.columns = existing_patients.columns.str.strip()
        existing_trials = load_file(trials_file)
        existing_trials.columns = existing_trials.columns.str.strip()
        
        available_studies = sorted(existing_trials["Study"].unique().tolist())
        
        # Get existing sites
        patient_origin_col = None
        possible_origin_cols = ['PatientSite', 'OriginSite', 'Practice', 'PatientPractice', 'HomeSite', 'Site']
        for col in possible_origin_cols:
            if col in existing_patients.columns:
                patient_origin_col = col
                break
        
        existing_sites = sorted(existing_patients[patient_origin_col].dropna().unique().tolist()) if patient_origin_col else ["Ashfields", "Kiltearn"]
        
        new_patient_id = st.text_input("Patient ID")
        new_study = st.selectbox("Study", options=available_studies)
        new_start_date = st.date_input("Start Date")
        
        if patient_origin_col:
            new_site = st.selectbox(f"{patient_origin_col}", options=existing_sites + ["Add New..."])
            if new_site == "Add New...":
                new_site = st.text_input("New Site Name")
        else:
            new_site = st.text_input("Patient Site")
        
        # Validation
        validation_errors = []
        if new_patient_id and new_patient_id in existing_patients["PatientID"].astype(str).values:
            validation_errors.append("Patient ID already exists")
        if new_start_date and new_start_date > date.today():
            validation_errors.append("Start date cannot be in future")
        
        if validation_errors:
            for error in validation_errors:
                st.error(error)
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Add Patient", disabled=bool(validation_errors), use_container_width=True):
                # Create new patient record
                new_patient_data = {
                    "PatientID": new_patient_id,
                    "Study": new_study,
                    "StartDate": new_start_date,
                }
                
                if patient_origin_col:
                    new_patient_data[patient_origin_col] = new_site
                else:
                    new_patient_data["PatientPractice"] = new_site
                
                # Add other columns
                for col in existing_patients.columns:
                    if col not in new_patient_data:
                        new_patient_data[col] = ""
                
                new_row_df = pd.DataFrame([new_patient_data])
                updated_patients_df = pd.concat([existing_patients, new_row_df], ignore_index=True)
                
                # Create download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    updated_patients_df.to_excel(writer, index=False, sheet_name="Patients")
                
                st.session_state.updated_patients_file = output.getvalue()
                st.session_state.updated_filename = f"Patients_Updated_{new_start_date.strftime('%Y%m%d')}.xlsx"
                st.session_state.patient_added = True
                st.session_state.show_patient_form = False
                
                st.success("Patient added successfully!")
                st.rerun()
        
        with col2:
            if st.button("Cancel", use_container_width=True):
                st.session_state.show_patient_form = False
                st.rerun()
    
    patient_form()

def extract_site_summary(patients_df, screen_failures=None):
    if patients_df.empty:
        return pd.DataFrame()
    
    unique_sites = sorted(patients_df["Site"].unique())
    site_summary_data = []
    
    for site in unique_sites:
        site_patients = patients_df[patients_df["Site"] == site]
        site_studies = site_patients["Study"].unique()
        
        site_screen_fails = 0
        if screen_failures:
            for _, patient in site_patients.iterrows():
                patient_study_key = f"{patient['PatientID']}_{patient['Study']}"
                if patient_study_key in screen_failures:
                    site_screen_fails += 1
        
        site_summary_data.append({
            "Site": site,
            "Patients": len(site_patients),
            "Screen Failures": site_screen_fails,
            "Active Patients": len(site_patients) - site_screen_fails,
            "Studies": ", ".join(sorted(site_studies))
        })
    
    return pd.DataFrame(site_summary_data)

def main():
    st.set_page_config(page_title="Clinical Trial Calendar Generator", layout="wide")
    st.title("🏥 Clinical Trial Calendar Generator")
    st.caption("v2.2.2 | Advanced Features with Modal Dialogs")

    # File uploaders
    st.sidebar.header("📁 Upload Data Files")
    patients_file = st.sidebar.file_uploader("Upload Patients File", type=['csv', 'xls', 'xlsx'])
    trials_file = st.sidebar.file_uploader("Upload Trials File", type=['csv', 'xls', 'xlsx'])
    actual_visits_file = st.sidebar.file_uploader("Upload Actual Visits File (Optional)", type=['csv', 'xls', 'xlsx'])
    
    # Store in session state
    st.session_state.patients_file = patients_file
    st.session_state.trials_file = trials_file
    st.session_state.actual_visits_file = actual_visits_file

    if patients_file and trials_file:
        # Action buttons
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            if st.button("➕ Add New Patient", use_container_width=True):
                st.session_state.show_patient_form = True
        with col2:
            if st.button("📋 Record Visit", use_container_width=True):
                st.session_state.show_visit_form = True
        
        # Modal dialogs
        if st.session_state.get('show_patient_form', False):
            try:
                patient_entry_modal()
            except AttributeError:
                st.error("Modal dialogs require Streamlit 1.28+")
                st.session_state.show_patient_form = False
        
        # Download buttons for added data
        if st.session_state.get('patient_added', False):
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    "💾 Download Updated Patients File",
                    data=st.session_state.updated_patients_file,
                    file_name=st.session_state.updated_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                if st.button("✅ Done", use_container_width=True):
                    st.session_state.patient_added = False
                    st.rerun()
            st.info("Patient added! Download and re-upload to see changes.")
            st.divider()

        try:
            # Load and process files
            patients_df = normalize_columns(load_file(patients_file))
            trials_df = normalize_columns(load_file(trials_file))
            actual_visits_df = normalize_columns(load_file(actual_visits_file)) if actual_visits_file else None

            # Date parsing
            patients_df, failed_patients = parse_dates_column(patients_df, "StartDate")
            if failed_patients:
                st.error(f"Unparseable StartDate values: {failed_patients}")

            if actual_visits_df is not None:
                actual_visits_df, failed_actuals = parse_dates_column(actual_visits_df, "ActualDate")
                if failed_actuals:
                    st.error(f"Unparseable ActualDate values: {failed_actuals}")

            # Data type conversion
            patients_df["PatientID"] = patients_df["PatientID"].astype(str)
            if "VisitNo" in trials_df.columns:
                trials_df["VisitNo"] = trials_df["VisitNo"].astype(str)
            if actual_visits_df is not None and "VisitNo" in actual_visits_df.columns:
                actual_visits_df["VisitNo"] = actual_visits_df["VisitNo"].astype(str)

            # Check missing studies
            missing_studies = set(patients_df["Study"].astype(str)) - set(trials_df["Study"].astype(str))
            if missing_studies:
                st.error(f"❌ Missing Study Definitions: {missing_studies}")
                st.stop()

            # Build calendar
            visits_df, calendar_df, stats, messages, site_column_mapping, unique_sites = build_calendar(
                patients_df, trials_df, actual_visits_df
            )
            
            # Extract screen failures
            screen_failures = {}
            if actual_visits_df is not None:
                screen_fail_visits = actual_visits_df[
                    actual_visits_df["Notes"].str.contains("ScreenFail", case=False, na=False)
                ]
                for _, visit in screen_fail_visits.iterrows():
                    patient_study_key = f"{visit['PatientID']}_{visit['Study']}"
                    screen_fail_date = visit['ActualDate']
                    if patient_study_key not in screen_failures or screen_fail_date < screen_failures[patient_study_key]:
                        screen_failures[patient_study_key] = screen_fail_date

            # Display processing messages
            if messages:
                with st.expander("📋 Processing Log", expanded=False):
                    for message in messages:
                        st.write(message)

            # Site summary
            site_summary_df = extract_site_summary(patients_df, screen_failures)
            if not site_summary_df.empty:
                st.subheader("Site Summary")
                st.dataframe(site_summary_df, use_container_width=True)

            # Main displays
            show_legend(actual_visits_df)
            display_styled_calendar(calendar_df, site_column_mapping, unique_sites)
            display_financial_analysis(stats, visits_df)
            
            # Quarterly profit sharing
            financial_df = visits_df[
                (visits_df['Visit'].str.startswith("✅")) |
                (visits_df['Visit'].str.startswith("❌ Screen Fail")) |
                (visits_df['Visit'].str.contains('Visit', na=False) & (~visits_df.get('IsActual', False)))
            ].copy()
            
            if not financial_df.empty:
                # Add time columns for profit sharing
                financial_df['Quarter'] = financial_df['Date'].dt.quarter
                financial_df['Year'] = financial_df['Date'].dt.year
                financial_df['QuarterYear'] = financial_df['Year'].astype(str) + '-Q' + financial_df['Quarter'].astype(str)
                
                display_quarterly_profit_sharing(financial_df, patients_df)

            # Site-wise breakdown by month with financial year summaries
            st.subheader("Site-wise Statistics by Month")
            
            # Add month-year and financial year columns to visits_df if not present
            visits_df['MonthYear'] = visits_df['Date'].dt.to_period('M')
            visits_df['FinancialYear'] = visits_df['Date'].apply(
                lambda d: f"{d.year}-{d.year+1}" if d.month >= 4 else f"{d.year-1}-{d.year}"
            )
            
            # Get unique months and financial years, sorted
            months = sorted(visits_df['MonthYear'].unique())
            financial_years = sorted(visits_df['FinancialYear'].unique())
            
            monthly_site_stats = []
            
            # Process monthly data
            for month in months:
                month_visits = visits_df[visits_df['MonthYear'] == month]
                month_patients = patients_df[patients_df['PatientID'].isin(month_visits['PatientID'].unique())]
                
                # Get financial year for this month
                fy = month_visits['FinancialYear'].iloc[0] if len(month_visits) > 0 else ""
                
                for site in unique_sites:
                    # Get patients from this site active in this month
                    site_patients = month_patients[month_patients["Site"] == site]
                    
                    # Get visits for patients from this site in this month
                    site_patient_ids = site_patients['PatientID'].unique()
                    site_visits = month_visits[month_visits["PatientID"].isin(site_patient_ids)]
                    
                    # Filter relevant visits (exclude tolerance periods)
                    relevant_visits = site_visits[
                        (site_visits["Visit"].str.startswith("✅")) | 
                        (site_visits["Visit"].str.startswith("❌ Screen Fail")) | 
                        (site_visits["Visit"].str.contains("Visit", na=False))
                    ]
                    
                    # Calculate metrics for this site and month
                    site_income = relevant_visits["Payment"].sum()
                    completed_visits = len(relevant_visits[relevant_visits["Visit"].str.startswith("✅")])
                    screen_fail_visits = len(relevant_visits[relevant_visits["Visit"].str.startswith("❌ Screen Fail")])
                    total_visits = len(relevant_visits)
                    pending_visits = total_visits - completed_visits - screen_fail_visits
                    
                    # Count new patients recruited this month
                    new_patients = len(site_patients)
                    
                    monthly_site_stats.append({
                        'Period': str(month),
                        'Financial Year': fy,
                        'Type': 'Month',
                        'Site': site,
                        'New Patients': new_patients,
                        'Completed Visits': completed_visits,
                        'Screen Fail Visits': screen_fail_visits,
                        'Pending Visits': pending_visits,
                        'Total Visits': total_visits,
                        'Income': f"£{site_income:,.2f}"
                    })
            
            # Add financial year summaries
            for fy in financial_years:
                fy_visits = visits_df[visits_df['FinancialYear'] == fy]
                fy_patients = patients_df[patients_df['PatientID'].isin(fy_visits['PatientID'].unique())]
                
                for site in unique_sites:
                    # Get patients from this site for this financial year
                    site_patients = fy_patients[fy_patients["Site"] == site]
                    
                    # Get visits for patients from this site in this financial year
                    site_patient_ids = site_patients['PatientID'].unique()
                    site_visits = fy_visits[fy_visits["PatientID"].isin(site_patient_ids)]
                    
                    # Filter relevant visits
                    relevant_visits = site_visits[
                        (site_visits["Visit"].str.startswith("✅")) | 
                        (site_visits["Visit"].str.startswith("❌ Screen Fail")) | 
                        (site_visits["Visit"].str.contains("Visit", na=False))
                    ]
                    
                    # Calculate annual metrics
                    site_income = relevant_visits["Payment"].sum()
                    completed_visits = len(relevant_visits[relevant_visits["Visit"].str.startswith("✅")])
                    screen_fail_visits = len(relevant_visits[relevant_visits["Visit"].str.startswith("❌ Screen Fail")])
                    total_visits = len(relevant_visits)
                    pending_visits = total_visits - completed_visits - screen_fail_visits
                    
                    # Count screen failures for this financial year
                    site_screen_fails = 0
                    for _, patient in site_patients.iterrows():
                        patient_study_key = f"{patient['PatientID']}_{patient['Study']}"
                        if patient_study_key in screen_failures:
                            screen_fail_date = screen_failures[patient_study_key]
                            # Check if screen failure occurred in this financial year
                            screen_fail_fy = f"{screen_fail_date.year}-{screen_fail_date.year+1}" if screen_fail_date.month >= 4 else f"{screen_fail_date.year-1}-{screen_fail_date.year}"
                            if screen_fail_fy == fy:
                                site_screen_fails += 1
                    
                    total_patients_fy = len(site_patients)
                    active_patients = total_patients_fy - site_screen_fails
                    
                    monthly_site_stats.append({
                        'Period': f"FY {fy}",
                        'Financial Year': fy,
                        'Type': 'Financial Year',
                        'Site': site,
                        'New Patients': f"{total_patients_fy} ({active_patients} active)",
                        'Completed Visits': completed_visits,
                        'Screen Fail Visits': screen_fail_visits,
                        'Pending Visits': pending_visits,
                        'Total Visits': total_visits,
                        'Income': f"£{site_income:,.2f}"
                    })
            
            if monthly_site_stats:
                # Sort by financial year, type, period, and site
                monthly_site_stats.sort(key=lambda x: (x['Financial Year'], x['Type'] == 'Financial Year', x['Period'], x['Site']))
                
                # Create separate tables for each site
                for site in unique_sites:
                    st.write(f"**{site} Practice**")
                    
                    site_data = [stat for stat in monthly_site_stats if stat['Site'] == site]
                    site_df = pd.DataFrame(site_data)
                    
                    # Drop the Site column since it's redundant in site-specific tables
                    display_df = site_df.drop('Site', axis=1)
                    
                    # Style the dataframe to highlight financial year rows
                    def highlight_fy_rows(row):
                        if row['Type'] == 'Financial Year':
                            return ['background-color: #e6f3ff; font-weight: bold'] * len(row)
                        else:
                            return [''] * len(row)
                    
                    styled_site_df = display_df.style.apply(highlight_fy_rows, axis=1)
                    st.dataframe(styled_site_df, use_container_width=True)
                    st.write("")  # Add space between sites

            # Monthly analysis by site
            st.subheader("Monthly Analysis by Site")
            
            # Filter only actual visits and main scheduled visits
            analysis_visits = visits_df[
                (visits_df['Visit'].str.startswith("✅")) |  # Actual completed visits
                (visits_df['Visit'].str.startswith("❌ Screen Fail")) |  # Screen failure visits
                (visits_df['Visit'].str.contains('Visit', na=False) & (~visits_df.get('IsActual', False)))  # Scheduled main visits
            ]
            
            if not analysis_visits.empty:
                # Analysis 1: Visits by Site of Visit (where visits happen)
                st.write("**Analysis by Visit Location (Where visits occur)**")
                visits_by_site_month = analysis_visits.groupby(['SiteofVisit', 'MonthYear']).size().reset_index(name='Visits')
                
                if not visits_by_site_month.empty:
                    visits_pivot = visits_by_site_month.pivot(index='MonthYear', columns='SiteofVisit', values='Visits').fillna(0)
                    
                    # Calculate visit ratios
                    visits_pivot['Total_Visits'] = visits_pivot.sum(axis=1)
                    visit_sites = [col for col in visits_pivot.columns if col != 'Total_Visits']
                    for site in visit_sites:
                        visits_pivot[f'{site}_Ratio'] = (visits_pivot[site] / visits_pivot['Total_Visits'] * 100).round(1)
                    
                    # Count unique patients by visit site per month
                    patients_by_visit_site_month = analysis_visits.groupby(['SiteofVisit', 'MonthYear'])['PatientID'].nunique().reset_index(name='Patients')
                    patients_visit_pivot = patients_by_visit_site_month.pivot(index='MonthYear', columns='SiteofVisit', values='Patients').fillna(0)
                    
                    # Calculate patient ratios for visit site
                    patients_visit_pivot['Total_Patients'] = patients_visit_pivot.sum(axis=1)
                    for site in visit_sites:
                        if site in patients_visit_pivot.columns:
                            patients_visit_pivot[f'{site}_Ratio'] = (patients_visit_pivot[site] / patients_visit_pivot['Total_Patients'] * 100).round(1)
                    
                    # Analysis 2: Patients by Origin Site (where patients come from)
                    st.write("**Analysis by Patient Origin (Where patients come from)**")
                    patients_by_origin_month = analysis_visits.groupby(['PatientOrigin', 'MonthYear'])['PatientID'].nunique().reset_index(name='Patients')
                    patients_origin_pivot = patients_by_origin_month.pivot(index='MonthYear', columns='PatientOrigin', values='Patients').fillna(0)
                    
                    # Calculate patient origin ratios
                    patients_origin_pivot['Total_Patients'] = patients_origin_pivot.sum(axis=1)
                    origin_sites = [col for col in patients_origin_pivot.columns if col != 'Total_Patients']
                    for site in origin_sites:
                        patients_origin_pivot[f'{site}_Ratio'] = (patients_origin_pivot[site] / patients_origin_pivot['Total_Patients'] * 100).round(1)
                    
                    # Display tables
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**Monthly Visits by Visit Site**")
                        visits_display = visits_pivot.copy()
                        visits_display.index = visits_display.index.astype(str)
                        
                        # Reorder columns
                        display_cols = []
                        for site in sorted(visit_sites):
                            display_cols.append(site)
                        for site in sorted(visit_sites):
                            ratio_col = f'{site}_Ratio'
                            if ratio_col in visits_display.columns:
                                display_cols.append(ratio_col)
                        display_cols.append('Total_Visits')
                        
                        visits_display = visits_display[display_cols]
                        st.dataframe(visits_display, use_container_width=True)
                    
                    with col2:
                        st.write("**Monthly Patients by Visit Site**")
                        patients_visit_display = patients_visit_pivot.copy()
                        patients_visit_display.index = patients_visit_display.index.astype(str)
                        
                        # Reorder columns
                        display_cols = []
                        for site in sorted(visit_sites):
                            if site in patients_visit_display.columns:
                                display_cols.append(site)
                        for site in sorted(visit_sites):
                            ratio_col = f'{site}_Ratio'
                            if ratio_col in patients_visit_display.columns:
                                display_cols.append(ratio_col)
                        display_cols.append('Total_Patients')
                        
                        patients_visit_display = patients_visit_display[display_cols]
                        st.dataframe(patients_visit_display, use_container_width=True)
                    
                    # Patient Origin Analysis
                    st.write("**Monthly Patients by Origin Site (Where patients come from)**")
                    patients_origin_display = patients_origin_pivot.copy()
                    patients_origin_display.index = patients_origin_display.index.astype(str)
                    
                    # Reorder columns
                    display_cols = []
                    for site in sorted(origin_sites):
                        display_cols.append(site)
                    for site in sorted(origin_sites):
                        ratio_col = f'{site}_Ratio'
                        if ratio_col in patients_origin_display.columns:
                            display_cols.append(ratio_col)
                    display_cols.append('Total_Patients')
                    
                    patients_origin_display = patients_origin_display[display_cols]
                    st.dataframe(patients_origin_display, use_container_width=True)
                    
                    # Cross-tabulation: Origin vs Visit Site
                    st.write("**Cross-Analysis: Patient Origin vs Visit Site**")
                    cross_tab = analysis_visits.groupby(['PatientOrigin', 'SiteofVisit'])['PatientID'].nunique().reset_index(name='Patients')
                    cross_pivot = cross_tab.pivot(index='PatientOrigin', columns='SiteofVisit', values='Patients').fillna(0)
                    cross_pivot['Total'] = cross_pivot.sum(axis=1)
                    
                    # Add row percentages
                    for col in cross_pivot.columns:
                        if col != 'Total':
                            cross_pivot[f'{col}_%'] = (cross_pivot[col] / cross_pivot['Total'] * 100).round(1)
                    
                    st.dataframe(cross_pivot, use_container_width=True)
                    
                    # Charts
                    st.subheader("Monthly Trends")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.write("**Visits by Visit Site**")
                        if not visits_pivot.empty:
                            chart_data = visits_pivot[[col for col in visits_pivot.columns if not col.endswith('_Ratio') and col != 'Total_Visits']]
                            chart_data.index = chart_data.index.astype(str)
                            st.bar_chart(chart_data)
                    
                    with col2:
                        st.write("**Patients by Visit Site**") 
                        if not patients_visit_pivot.empty:
                            chart_data = patients_visit_pivot[[col for col in patients_visit_pivot.columns if not col.endswith('_Ratio') and col != 'Total_Patients']]
                            chart_data.index = chart_data.index.astype(str)
                            st.bar_chart(chart_data)
                    
                    with col3:
                        st.write("**Patients by Origin Site**")
                        if not patients_origin_pivot.empty:
                            chart_data = patients_origin_pivot[[col for col in patients_origin_pivot.columns if not col.endswith('_Ratio') and col != 'Total_Patients']]
                            chart_data.index = chart_data.index.astype(str)
                            st.bar_chart(chart_data)

            # Enhanced downloads
            st.subheader("💾 Download Options")
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter
                
                final_ordered_columns = ["Date", "Day"]
                for site in unique_sites:
                    site_columns = site_column_mapping.get(site, [])
                    for col in site_columns:
                        if col in calendar_df.columns:
                            final_ordered_columns.append(col)
                
                excel_df = calendar_df[final_ordered_columns].copy()
                excel_df["Date"] = excel_df["Date"].dt.strftime("%d/%m/%Y")
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    excel_df.to_excel(writer, index=False, sheet_name="VisitCalendar", startrow=1)
                    ws = writer.sheets["VisitCalendar"]

                    # Site headers
                    for col_idx, col_name in enumerate(excel_df.columns, 1):
                        col_letter = get_column_letter(col_idx)
                        if col_name not in ["Date", "Day"]:
                            for site in unique_sites:
                                if col_name in site_column_mapping.get(site, []):
                                    ws[f"{col_letter}1"] = site
                                    ws[f"{col_letter}1"].font = Font(bold=True, size=12)
                                    ws[f"{col_letter}1"].fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
                                    ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")
                                    break

                    # Auto-adjust column widths
                    for idx, col in enumerate(excel_df.columns, 1):
                        col_letter = get_column_letter(idx)
                        max_length = max([len(str(cell)) for cell in excel_df[col].tolist()] + [len(col)])
                        ws.column_dimensions[col_letter].width = max(10, max_length + 2)

                st.download_button(
                    "📊 Excel with Site Headers & Full Formatting",
                    data=output.getvalue(),
                    file_name="VisitCalendar_Complete.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except ImportError:
                st.warning("Excel formatting unavailable - install openpyxl for enhanced features")
                buf = io.BytesIO()
                calendar_df.to_excel(buf, index=False)
                st.download_button("💾 Download Basic Excel", data=buf.getvalue(), file_name="VisitCalendar.xlsx")
                
        except Exception as e:
            st.error(f"Error processing files: {e}")
            st.exception(e)
    
    else:
        st.info("Please upload both Patients and Trials files to get started.")
        
        st.markdown("""
        ### Expected File Structure:
        
        **Patients File:**
        - PatientID, Study, StartDate
        - Site/PatientPractice (optional - for patient origin)
        
        **Trials File:**
        - Study, Day, VisitNo, SiteforVisit
        - Income/Payment, ToleranceBefore, ToleranceAfter (optional)
        
        **Actual Visits File (Optional):**
        - PatientID, Study, VisitNo, ActualDate
        - ActualPayment, Notes (optional)
        - Use 'ScreenFail' in Notes to stop future visits
        """)

if __name__ == "__main__":
    main()
